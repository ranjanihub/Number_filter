import streamlit as st
import pandas as pd
import time
import re
import openpyxl
import io

# Configure the Streamlit page
st.set_page_config(
    page_title="Phone Number Filter", 
    page_icon="", 
    layout="wide"
)

def is_empty(val):
    """Check if a cell value is considered empty."""
    if pd.isna(val) or val is None:
        return True
    if str(val).strip() == '':
        return True
    return False

def process_number(num_str):
    """
    Extracts and validates a 10-digit mobile number.
    Returns the 10-digit number or None if invalid.
    """
    num_str = str(num_str)
    
    # Remove all non-digit characters (like spaces, dashes, etc.)
    digits = re.sub(r'\D', '', num_str)
    
    # If the number is 11 digits and starts with 0, remove the leading 0
    if len(digits) == 11 and digits.startswith('0'):
        digits = digits[1:]
        
    # If the number is 12 digits and starts with India's country code (91), remove it
    elif len(digits) == 12 and digits.startswith('91'):
        digits = digits[2:]
    
    # Valid mobile numbers must be EXACTLY 10 digits and start with 6, 7, 8, or 9
    if len(digits) == 10 and digits[0] in '6789':
        return digits
        
    return None

def process_csv_chunked(file, remove_duplicates, progress_text, valid_numbers=None):
    """
    Memory optimized CSV processing using pandas chunk_size.
    We iterate over chunks of data instead of loading the entire file.
    """
    if valid_numbers is None:
        valid_numbers = set()
    total_rows = 0
    total_valid = 0
    total_invalid = 0
    
    output_rows = []
    headers = ["Phone Number"]
    
    # Using low_memory=False to prevent mixed type warnings and chunksize to avoid memory overload
    chunk_iter = pd.read_csv(file, chunksize=50000, dtype=str, on_bad_lines='skip', low_memory=False)
    
    for chunk in chunk_iter:
        total_rows += len(chunk)
        
        # Update progress text
        progress_text.text(f"Processing... {total_rows:,} rows scanned.")
        
        for row_tuple in chunk.itertuples(index=False, name=None):
            extracted_numbers = []
            row_list = list(row_tuple)
            
            for i, val in enumerate(row_list):
                if is_empty(val):
                    continue
                
                parts = re.split(r'[,/\|\n]', str(val))
                for pt in parts:
                    cleaned = process_number(pt)
                    if cleaned:
                        extracted_numbers.append(cleaned)
            
            if extracted_numbers:
                for num in extracted_numbers:
                    if remove_duplicates:
                        if num not in valid_numbers:
                            valid_numbers.add(num)
                            output_rows.append([num])
                            total_valid += 1
                    else:
                        output_rows.append([num])
                        total_valid += 1
            else:
                total_invalid += 1
                
    return output_rows, headers, total_rows, total_valid, total_invalid

def process_excel_iterative(file, remove_duplicates, progress_text, selected_sheets=None, valid_numbers=None):
    """
    Memory optimized Excel processing using openpyxl in read-only mode.
    Generator-based iteration prevents holding entire rows/sheets in memory.
    """
    if valid_numbers is None:
        valid_numbers = set()
    total_rows = 0
    total_valid = 0
    total_invalid = 0
    
    output_rows = []
    headers = ["Phone Number"]
    
    try:
        # read_only=True ensures memory efficient row-by-row iteration without loading whole document
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        
        if selected_sheets:
            sheets_to_process = [s for s in selected_sheets if s in wb.sheetnames]
        else:
            sheets_to_process = wb.sheetnames

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            
            # Use values_only=True to get raw values directly, skipping cell objects metadata (much faster/lighter)
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                total_rows += 1
                
                # Update UI periodically to prevent slowing down the loop
                if total_rows % 5000 == 0:
                    progress_text.text(f"Processing... {total_rows:,} rows scanned.")
                    
                extracted_numbers = []
                # Ensure row is a list
                row_list = list(row) if row is not None else []
                
                for i, val in enumerate(row_list):
                    if is_empty(val):
                        continue
                        
                    parts = re.split(r'[,/\|\n]', str(val))
                    for pt in parts:
                        cleaned = process_number(pt)
                        if cleaned:
                            extracted_numbers.append(cleaned)
                            
                if extracted_numbers:
                    for num in extracted_numbers:
                        if remove_duplicates:
                            if num not in valid_numbers:
                                valid_numbers.add(num)
                                output_rows.append([num])
                                total_valid += 1
                        else:
                            output_rows.append([num])
                            total_valid += 1
                else:
                    total_invalid += 1
        wb.close()
    except Exception as e:
        raise RuntimeError(f"Failed to read Excel file appropriately. Make sure the file isn't corrupted. Error details: {str(e)}")
        
    return output_rows, headers, total_rows, total_valid, total_invalid

def main():
    st.title("Indian Mobile Number Extractor")
    st.markdown("Upload large CSV or Excel files (up to **1GB**) to extract valid Indian mobile phone numbers efficiently.")
    
    # Instructions for 1GB limit configuration
    st.sidebar.markdown("### Upload Instructions")
    st.sidebar.info(
        "**Note:**\n\n"
        "If **CSV File** is not uploaded, then upload the **Excel file**.\n"
        "\nStreamlit upload limit has been increased to 1GB."
    )
    
    st.sidebar.header("Processing Settings")
    remove_duplicates = st.sidebar.checkbox("Remove Duplicate Numbers", value=True, help="Toggle to filter out identical numbers across all columns.")
    
    # File uploader
    uploaded_files = st.file_uploader("Choose CSV or Excel files", type=['csv', 'xlsx'], accept_multiple_files=True)
    
    if uploaded_files:
        st.write(f"**Total Files Uploaded:** {len(uploaded_files)}")
        
        # If there's exactly one Excel file, allow sheet selection
        excel_files = [f for f in uploaded_files if not f.name.lower().endswith('.csv')]
        selected_sheets = None
        
        if len(excel_files) == 1:
            try:
                # Load workbook efficiently to fetch sheet names
                wb_temp = openpyxl.load_workbook(excel_files[0], read_only=True)
                wb_sheets = wb_temp.sheetnames
                wb_temp.close()
                selected_sheets = st.multiselect(f"Select Sheets for {excel_files[0].name}", wb_sheets, default=wb_sheets)
                excel_files[0].seek(0)
                
                if not selected_sheets:
                    st.warning("Please select at least one sheet to process.")
                    return
            except Exception as e:
                st.error(f"Error reading Excel file sheets: {e}")
                return
        elif len(excel_files) > 1:
            st.info("Multiple Excel files detected. All sheets in all Excel files will be processed.")

        # UI controls and processing triggers
        col_btn, _ = st.columns([1, 4])
        with col_btn:
            start_process = st.button("Process All Files", use_container_width=True)
            
        if start_process:
            start_time = time.time()
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            try:
                all_output_rows = []
                all_total_rows = 0
                all_total_valid = 0
                all_total_invalid = 0
                total_valid_numbers_set = set() # For cross-file deduplication
                
                num_files = len(uploaded_files)
                
                for i, uploaded_file in enumerate(uploaded_files):
                    file_name = uploaded_file.name
                    file_size_mb = uploaded_file.size / (1024 * 1024)
                    
                    if file_size_mb > 1024:
                        st.error(f"Skipping {file_name}: File size ({file_size_mb:.2f}MB) exceeds 1GB limit.")
                        continue
                    
                    progress_text.text(f"File {i+1}/{num_files}: {file_name}...")
                    
                    is_csv = file_name.lower().endswith('.csv')
                    
                    if is_csv:
                        results = process_csv_chunked(uploaded_file, remove_duplicates, progress_text, valid_numbers=total_valid_numbers_set)
                    else:
                        sheets = selected_sheets if (len(excel_files) == 1) else None
                        results = process_excel_iterative(uploaded_file, remove_duplicates, progress_text, sheets, valid_numbers=total_valid_numbers_set)
                    
                    output_rows, headers, total_rows_proc, total_valid, total_invalid = results
                    all_output_rows.extend(output_rows)
                    all_total_rows += total_rows_proc
                    all_total_valid += total_valid
                    all_total_invalid += total_invalid
                    
                    # Update global progress bar based on files count
                    progress_bar.progress(int((i + 1) / num_files * 100))
                
                processing_time = time.time() - start_time
                progress_text.empty()
                st.success(f"Processing Complete in {processing_time:.2f} seconds!")
                
                # Render Metrics
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Rows Processed", f"{all_total_rows:,}")
                col2.metric("Valid Numbers", f"{all_total_valid:,}")
                col3.metric("Invalid Rows/Cells", f"{all_total_invalid:,}")
                
                valid_percent = 0
                if (all_total_valid + all_total_invalid) > 0:
                    valid_percent = (all_total_valid / (all_total_valid + all_total_invalid)) * 100
                col4.metric("Valid Ratio", f"{valid_percent:.2f}%")
                
                if len(all_output_rows) > 0:
                    df_out = pd.DataFrame(all_output_rows, columns=["Phone Number"])
                    
                    st.divider()
                    st.subheader("Preview (First 100 Entries)")
                    st.dataframe(df_out.head(100), use_container_width=True)
                    
                    # Store valid list internally and generate bytes for UI
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_out.to_excel(writer, index=False, sheet_name='Filtered Numbers')
                    
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="Download Combined Cleaned File as Excel",
                        data=excel_data,
                        file_name=f'Combined_Phone_Numbers_{int(time.time())}.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        type="primary"
                    )
                else:
                    st.warning("No valid Indian phone numbers were found in the uploaded files.")
                    
            except Exception as e:
                progress_bar.empty()
                progress_text.empty()
                st.error(f"An error occurred during processing:\n{str(e)}")

if __name__ == "__main__":
    main()
